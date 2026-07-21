from apps.core.models import AuditLog
from apps.core.middleware import get_current_user, get_current_ip

def log_action(action_type, instance, changes):
    """
    سرویس ثبت خودکار لاگ تغییرات با استفاده از کاربر و آی‌پی جاری سشن.
    """
    user = get_current_user()
    ip = get_current_ip()
    model_name = instance._meta.model_name
    object_id = str(instance.pk)

    log = AuditLog(
        user=user,
        action_type=action_type,
        model_name=model_name,
        object_id=object_id,
        changes=changes,
        ip_address=ip
    )
    log.save()


def export_to_excel_response(filename, headers, rows):
    """
    تولید فایل اکسل راست‌چین شده با هدرهای زیبا و خروجی مستقیم به پاسخ HTTP
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse
    from urllib.parse import quote

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "گزارش"
    
    # تنظیم راست‌چین بودن صفحه برای زبان فارسی
    ws.views.sheetView[0].showGridLines = True
    ws.sheet_properties.tabColor = "1E293B"
    ws.sheet_view.rightToLeft = True

    # تنظیمات استایل‌دهی
    font_name = "Tahoma"  # فونت استاندارد سیستم برای فایل‌های اکسل فارسی
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name=font_name, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=font_name, size=10)
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center', wrap_text=True)

    # نوشتن هدرها
    ws.append(headers)
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
    
    # نوشتن ردیف‌های داده
    for row_num, row_data in enumerate(rows, start=2):
        ws.append(row_data)
        for col_num in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.font = data_font
            cell.alignment = right_align
            cell.border = thin_border

    # تنظیم خودکار عرض ستون‌ها
    for col in ws.columns:
        max_len = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or '')
            # تقریب طول رشته فارسی به بایت
            val_len = len(val.encode('utf-8'))
            if val_len > max_len:
                max_len = val_len
        ws.column_dimensions[col_letter].width = min(max(max_len // 2 + 4, 12), 45)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f"attachment; filename*=utf-8''{quote(filename)}"
    wb.save(response)
    return response
