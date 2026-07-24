"""
سرویس‌های import امن اسکرینینگ - Zero-Risk Protocol
فاز ۱: Dry-Run (بدون تغییر DB)
فاز ۲: Commit (تراکنش اتمیک)
"""
import openpyxl
import logging

logger = logging.getLogger(__name__)


def dry_run_import_screening(file_path):
    """
    Dry-Run: اعتبارسنجی فایل اکسل غربالگری بدون تغییر دیتابیس.
    
    ستون‌های مورد انتظار:
        ExamCode, NationCode, Result, EXP
    
    Result values:
        مجاز, غیرمجاز, لغو درخواست, در انتظار غربالگری
    
    Returns:
        dict: {
            'total_rows': int,
            'valid_rows': list of dict,
            'invalid_rows': list of dict (with error reason),
            'summary': dict of counts
        }
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    # تشخیص ستون‌ها از هدر
    headers = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]
    
    # مپ ستون‌ها
    col_map = {}
    for idx, h in enumerate(headers):
        if 'examcode' in h.lower() or 'exam' in h.lower() or 'کد' in h:
            col_map['exam_code'] = idx
        elif 'nationcode' in h.lower() or 'nation' in h.lower() or 'ملی' in h.lower():
            col_map['nation_code'] = idx
        elif 'result' in h.lower() or 'نتیجه' in h:
            col_map['result'] = idx
        elif h in ('EXP', 'exp', 'علت', 'دلیل', 'توضیح'):
            col_map['reason'] = idx
    
    # اگر مپ خودکار نشد، از مپ دستی استفاده کن
    if 'exam_code' not in col_map and len(headers) >= 2:
        col_map['exam_code'] = 1  # ستون دوم: ExamCode
    if 'nation_code' not in col_map and len(headers) >= 3:
        col_map['nation_code'] = 2  # ستون سوم: NationCode
    if 'result' not in col_map and len(headers) >= 4:
        col_map['result'] = 3  # ستون چهارم: Result
    if 'reason' not in col_map and len(headers) >= 5:
        col_map['reason'] = 4  # ستون پنجم: EXP
    # اگر دلیل در ستون ۶ هست
    if 'reason' not in col_map and len(headers) >= 6:
        col_map['reason'] = 5

    valid_rows = []
    invalid_rows = []
    counts = {'مجاز': 0, 'غیرمجاز': 0, 'لغو درخواست': 0, 'در انتظار غربالگری': 0}

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row:
            continue

        exam_code = str(row[col_map['exam_code']]).strip() if col_map.get('exam_code') is not None and row[col_map['exam_code']] else None
        nation_code = str(row[col_map['nation_code']]).strip() if col_map.get('nation_code') is not None and row[col_map['nation_code']] else None
        result = str(row[col_map['result']]).strip() if col_map.get('result') is not None and row[col_map['result']] else None
        reason = str(row[col_map['reason']]).strip() if col_map.get('reason') is not None and row[col_map['reason']] else None

        errors = []

        if not exam_code:
            errors.append('ExamCode خالی است')
        if not nation_code:
            errors.append('NationCode خالی است')
        if not result:
            errors.append('Result خالی است')
        elif result not in ('مجاز', 'غیرمجاز', 'لغو درخواست', 'در انتظار غربالگری'):
            errors.append(f'Result نامعتبر: {result}')

        if errors:
            invalid_rows.append({
                'row': row_idx,
                'exam_code': exam_code,
                'nation_code': nation_code,
                'result': result,
                'reason': reason,
                'errors': errors,
            })
            continue

        # اعتبارسنجی وجود در دیتابیس (بدون save)
        from apps.jobs.models import JobOpportunity
        from apps.candidates.models import Candidate, JobApplication

        db_errors = []

        job = JobOpportunity.objects.filter(code=exam_code, is_deleted=False).first()
        if not job:
            db_errors.append(f'فرصت شغلی با کد {exam_code} یافت نشد')

        candidate = Candidate.objects.filter(national_id=nation_code, is_deleted=False).first()
        if not candidate:
            db_errors.append(f'متقاضی با کد ملی {nation_code} یافت نشد')

        application = None
        if job and candidate:
            application = JobApplication.objects.filter(
                job=job, candidate=candidate, is_deleted=False
            ).first()
            if not application:
                db_errors.append(f'درخواست همکاری برای کد {exam_code} و کد ملی {nation_code} یافت نشد')

        if db_errors:
            invalid_rows.append({
                'row': row_idx,
                'exam_code': exam_code,
                'nation_code': nation_code,
                'result': result,
                'reason': reason,
                'errors': db_errors,
            })
            continue

        valid_rows.append({
            'row': row_idx,
            'exam_code': exam_code,
            'nation_code': nation_code,
            'result': result,
            'reason': reason,
            'application_id': application.id,
        })
        counts[result] = counts.get(result, 0) + 1

    wb.close()

    return {
        'total_rows': len(valid_rows) + len(invalid_rows),
        'valid_rows': valid_rows,
        'invalid_rows': invalid_rows,
        'summary': {
            'valid_count': len(valid_rows),
            'invalid_count': len(invalid_rows),
            'by_result': counts,
        },
    }


def commit_import_screening(valid_rows):
    """
    Commit: اعمال نتایج غربالگری روی JobApplication ها با تراکنش اتمیک.
    
    Args:
        valid_rows: list of dict (خروجی dry_run)
    
    Returns:
        dict: {'updated': int, 'failed': int, 'errors': list}
    """
    from django.db import transaction
    from apps.candidates.models import JobApplication

    updated = 0
    failed = 0
    errors = []

    try:
        with transaction.atomic():
            for row in valid_rows:
                try:
                    app = JobApplication.objects.select_for_update().get(
                        id=row['application_id'],
                        is_deleted=False,
                    )
                    app.screening_result = row['result']
                    app.screening_reason = row['reason'] or ''
                    app.save(update_fields=['screening_result', 'screening_reason'])
                    updated += 1
                except JobApplication.DoesNotExist:
                    failed += 1
                    errors.append(f"Application {row.get('application_id')} not found: row {row.get('row')}")
                except Exception as e:
                    failed += 1
                    errors.append(f"Row {row.get('row')}: {str(e)}")
    except Exception as e:
        logger.error(f"Transaction rolled back: {e}")
        return {'updated': 0, 'failed': len(valid_rows), 'errors': [str(e)]}

    return {
        'updated': updated,
        'failed': failed,
        'errors': errors,
    }
