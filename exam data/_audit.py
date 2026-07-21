import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import pandas as pd
import numpy as np
from openpyxl import load_workbook

f = r'D:\Payjoo-ATS-Demo\exam data\_main.xlsx'
f4101 = r'D:\Payjoo-ATS-Demo\exam data\4101.xlsx'

print('=' * 70)
print('1. مغایرت جدول وضعیت با داده واقعی')
print('=' * 70)

df_status = pd.read_excel(f, sheet_name='جدول وضعیت')
mismatches = 0
for _, r in df_status.iterrows():
    code = r['کد']
    if pd.isna(code):
        continue
    code = int(code)

    s_reg = r['تعداد ثبت‌نام']
    s_elg = r['تعداد واجد شرایط']
    s_written = r['تعداد حاضرین آزمون کتبی']
    s_skill_out = r['تعداد نفرات خروجی کتبی']
    s_interview = r['تعداد حاضرین مصاحبه']
    s_kanun = r['تعداد معرفی به کانون']

    df_reg = pd.read_excel(f, sheet_name='ثبت نام')
    a_reg = len(df_reg[df_reg['ExamCode'] == code])
    
    df_kot = pd.read_excel(f, sheet_name='کتبی')
    a_kot = len(df_kot[df_kot['ExamCode'] == code])
    
    df_mah = pd.read_excel(f, sheet_name='مهارتی')
    a_mah = len(df_mah[df_mah['ExamCode'] == code])
    
    df_int = pd.read_excel(f, sheet_name='مصاحبه')
    a_int = len(df_int[df_int['ExamCode'] == code])
    
    df_ac = pd.read_excel(f, sheet_name='کانون')
    a_ac = len(df_ac[df_ac['ExamCode'] == code])

    issues = []
    if not pd.isna(s_reg) and a_reg != s_reg and str(s_reg).strip('*').isdigit():
        issues.append(f'ثبت‌نام جدول={int(s_reg)} واقعی={a_reg}')
    if not pd.isna(s_elg) and a_kot != s_elg and str(s_elg).strip('*').isdigit():
        issues.append(f'واجدشرایط جدول={int(s_elg)} واقعی={a_kot}')
    if not pd.isna(s_skill_out) and a_mah != s_skill_out and str(s_skill_out).strip('*').isdigit():
        issues.append(f'خروجی کتبی جدول={int(s_skill_out)} واقعی={a_mah}')
    if not pd.isna(s_interview) and a_int != s_interview and str(s_interview).strip('*').isdigit():
        issues.append(f'مصاحبه جدول={int(s_interview)} واقعی={a_int}')
    if not pd.isna(s_kanun) and a_ac != s_kanun and str(s_kanun).strip('*').isdigit():
        issues.append(f'کانون جدول={int(s_kanun)} واقعی={a_ac}')

    if issues:
        mismatches += 1
        nm = r.get('عنوان پست', '')
        print(f'  کد {code} ({str(nm)[:30]}):')
        for iss in issues:
            print(f'    -> {iss}')

print(f'\nمجموع کدهای دارای مغایرت: {mismatches}')
print()

print('=' * 70)
print('2. بررسی شیت قبولی نهایی (ردیف‌های خالی)')
print('=' * 70)

wb = load_workbook(f, read_only=True, data_only=True)
ws = wb['قبولی نهایی']
non_empty = 0
total_rows = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    total_rows += 1
    if row[0] is not None:
        non_empty += 1
    if non_empty > 50 and total_rows > 1000:
        break
wb.close()
print(f'  ردیف‌های غیرخالی در 1000 ردیف اول: {non_empty}')
print(f'  از حداکثر 1,048,575 ردیف، اکثراً خالی هستند.')
print()

print('=' * 70)
print('3. مقایسه تعداد ثبت‌نام در فایل اصلی vs فایل‌های شماره‌دار')
print('=' * 70)

df_reg_all = pd.read_excel(f, sheet_name='ثبت نام')
print(f'  مجموع ثبت‌نام در فایل اصلی: {len(df_reg_all)} نفر')
exam_counts = df_reg_all['ExamCode'].value_counts()
print(f'  تعداد کدهای آزمون: {len(exam_counts)}')

# Check a few files
import os
exam_dir = r'D:\Payjoo-ATS-Demo\exam data'
files = sorted([x for x in os.listdir(exam_dir) if x.endswith('.xlsx') and x[:4].isdigit()])
print(f'  تعداد فایل‌های شماره‌دار: {len(files)}')
print()

print('=' * 70)
print('4. مغایرت مشخص در 4101 (جزئیات بیشتر)')
print('=' * 70)

# Check duplicate NationCodes in main file for exam 4101
df_kot_4101 = df_kot[df_kot['ExamCode'] == 4101]
dups = df_kot_4101[df_kot_4101.duplicated(subset=['NationCode'], keep=False)]
print(f'  کتبی 4101 - رکوردهای تکراری کد ملی: {len(dups)}')
if len(dups) > 0:
    print('  موارد تکراری:')
    for _, r in dups.iterrows():
        print(f'    {r["NationCode"]}')

# Check منفی بودن نمرات
neg = df_kot[df_kot['ScoreW'] < 0]
print(f'  نمرات منفی کتبی: {len(neg)}')

# Check if any registration has no progression
df_reg = pd.read_excel(f, sheet_name='ثبت نام')
df_kot = pd.read_excel(f, sheet_name='کتبی')
codes_without_kot = set(df_reg['ExamCode'].unique()) - set(df_kot['ExamCode'].unique())
if codes_without_kot:
    print(f'  کدهای ثبت‌نام بدون آزمون کتبی: {len(codes_without_kot)}')
else:
    print(f'  همه کدهای ثبت‌نام آزمون کتبی دارند.')

print()

print('=' * 70)
print('5. Column1, Column2, Column3 در جدول وضعیت')
print('=' * 70)

extra_cols = df_status[['کد', 'Column1', 'Column2', 'Column3']].dropna(how='all', subset=['Column1', 'Column2', 'Column3'])
if len(extra_cols) > 0:
    print(f'  تعداد ردیف‌های دارای داده در ستون‌های اضافی: {len(extra_cols)}')
    print(extra_cols.to_string())
else:
    print('  ستون‌های اضافی همگی خالی هستند.')

print()

print('=' * 70)
print('6. تعداد کل فایل‌ها vs تعداد کدهای آزمون')
print('=' * 70)
print(f'  تعداد فایل‌های شماره‌دار: {len(files)}')
print(f'  تعداد کدهای آزمون در جدول وضعیت: {df_status["کد"].nunique()}')
print(f'  تعداد کدهای آزمون در ثبت‌نام: {df_reg["ExamCode"].nunique()}')
print()

print('=' * 70)
print('7. بررسی داده‌های 4101 - تفاوت ثبت‌نام')
print('=' * 70)
df_r_4101 = pd.read_excel(f4101, sheet_name='ثبت نام')
print(f'  4101.xlsx ثبت‌نام: {len(df_r_4101)} نفر')
print(f'  فایل اصلی ثبت‌نام 4101: {a_reg} نفر')

# Check if there are entries in 4101.xlsx that are not in main file
# We need to match by something - but 4101.xlsx has personnel id, not national code in registration
# Let's check if the 2 extra are real people or duplicates
personnel_ids = df_r_4101['شماره پرسنلي'].dropna()
dups = df_r_4101[df_r_4101.duplicated(subset=['شماره پرسنلي'], keep=False)]
print(f'  افراد با شماره پرسنلی تکراری: {len(dups)}')
print(f'  تعداد شماره پرسنلی یکتا: {personnel_ids.nunique()}')
print()

print('=' * 70)
print('8. بررسی وضعیت افراد - عدم تطابق Result بین مراحل')
print('=' * 70)
df_kot = pd.read_excel(f, sheet_name='کتبی')
df_mah = pd.read_excel(f, sheet_name='مهارتی')
kot_ok = set(df_kot[df_kot['Result1'] == 'مجاز']['NationCode'].astype('Int64').astype(str))
mah_all = set(df_mah['NationCode'].astype('Int64').astype(str))
# People in مهارتی but NOT مجاز in written
not_ok = mah_all - kot_ok
print(f'  افرادی که به مرحله مهارتی رفته‌اند ولی در کتبی مجاز نبوده‌اند: {len(not_ok)}')

print('\nتحلیل کامل شد.')
