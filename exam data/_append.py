import openpyxl, os, shutil
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

path = r'D:\Payjoo-ATS-Demo\exam data'
src = os.path.join(path, 'ورود اطلاعات.xlsx')
out = os.path.join(path, 'ورود اطلاعات.xlsx')

# Backup
backup = src.replace('.xlsx', '_backup.xlsx')
shutil.copy2(src, backup)
print(f'Backup saved: {backup}')

# Load source
wb = openpyxl.load_workbook(src)
ws = wb['ثبت نام']

# Find last row
last_row = ws.max_row
print(f'Last row in file: {last_row} (header row 1 + {last_row-1} data rows)')

# Read existing keys to avoid duplicates
existing_keys = set()
for r in range(2, last_row + 1):
    key_val = ws.cell(row=r, column=6).value
    if key_val is not None:
        existing_keys.add(str(key_val))

print(f'Existing keys: {len(existing_keys)}')

# Get last row number from column A
last_row_num = 0
for r in range(2, last_row + 1):
    v = ws.cell(row=r, column=1).value
    if v is not None:
        try:
            last_row_num = max(last_row_num, int(v))
        except:
            pass

print(f'Last sequential row number: {last_row_num}')

# Missing exam codes
missing = ['4102','4177','4179','4180','4181','4182','4183','4184','4185','4186','4187','4188','4189','4190','4191','4192','4193','4194','4195','4196','4197','4198','4199','50013135','50027440','50038035','50047240','50088030']

# Define a thin border style
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

new_row_num = last_row_num
added = 0
skipped = 0

for code in missing:
    fp = os.path.join(path, f'{code}.xlsx')
    if not os.path.exists(fp):
        fp = os.path.join(path, f'{code}.XLSX')
    
    wb2 = openpyxl.load_workbook(fp, read_only=True, data_only=True)
    ws2 = wb2['ثبت نام']
    
    for row in ws2.iter_rows(min_row=2, values_only=True):
        if not any(v is not None for v in row):
            continue
        
        person_id = str(row[0]).strip() if row[0] is not None else ''
        if not person_id:
            continue
        
        key = code + person_id
        
        if key in existing_keys:
            skipped += 1
            continue
        
        existing_keys.add(key)
        new_row_num += 1
        last_row += 1
        
        # Write row
        ws.cell(row=last_row, column=1, value=new_row_num)
        ws.cell(row=last_row, column=2, value=int(code))
        ws.cell(row=last_row, column=3, value=int(person_id))
        ws.cell(row=last_row, column=4, value='مجاز')
        ws.cell(row=last_row, column=5, value=None)
        ws.cell(row=last_row, column=6, value=key)
        
        # Apply border like existing cells
        for c in range(1, 7):
            ws.cell(row=last_row, column=c).border = thin_border
        
        added += 1
    
    wb2.close()
    print(f'  {code}: added so far...')

print(f'\nAdded: {added}')
print(f'Skipped (duplicate key): {skipped}')
print(f'New total records: {new_row_num}')

wb.save(src)
print(f'\nSaved: {src}')
print(f'Total rows in sheet (header+data): {last_row}')
