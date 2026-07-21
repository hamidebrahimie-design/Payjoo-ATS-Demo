import openpyxl, os

path = r'D:\Payjoo-ATS-Demo\exam data'
missing = ['4102','4177','4179','4180','4181','4182','4183','4184','4185','4186','4187','4188','4189','4190','4191','4192','4193','4194','4195','4196','4197','4198','4199','50013135','50027440','50038035','50047240','50088030']

# 1. Load existing nationcodes
wb = openpyxl.load_workbook(os.path.join(path, 'ورود اطلاعات.xlsx'), read_only=True, data_only=True)
ws = wb['ثبت نام']
existing_nationcodes = set()
existing_total = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    if any(v is not None for v in row):
        existing_total += 1
        if row[2] is not None:
            existing_nationcodes.add(str(row[2]))
wb.close()

print(f'Existing in registration: {existing_total} records, {len(existing_nationcodes)} unique nationcodes')
print()

# 2. Extract from missing exam files
new_total = 0
new_unique_nc = set()
file_details = []

for code in missing:
    fp = os.path.join(path, f'{code}.xlsx')
    if not os.path.exists(fp):
        fp = os.path.join(path, f'{code}.XLSX')
    if not os.path.exists(fp):
        file_details.append((code, 0, 0, 0))
        continue
    
    wb2 = openpyxl.load_workbook(fp, read_only=True, data_only=True)
    ws2 = wb2['ثبت نام']
    cnt = 0
    nationcodes = set()
    for row in ws2.iter_rows(min_row=2, values_only=True):
        if any(v is not None for v in row):
            cnt += 1
            if row[0] is not None:
                nationcodes.add(str(row[0]))
    wb2.close()
    
    new_nc = nationcodes - existing_nationcodes
    file_details.append((code, cnt, len(nationcodes), len(new_nc)))
    new_total += cnt
    new_unique_nc.update(nationcodes)

# 3. Check overlap
overlap = len(new_unique_nc & existing_nationcodes)
new_unique_only = len(new_unique_nc - existing_nationcodes)

print('--- Breakdown of 28 files ---')
print(f'{"Exam":<12} {"Total":<8} {"Unique NC":<10} {"New NC":<8}')
print('-' * 42)
for code, cnt, uni, nnew in file_details:
    print(f'{code:<12} {cnt:<8} {uni:<10} {nnew:<8}')

print()
print('--- Summary ---')
print(f'Current records in main file: {existing_total}')
print(f'Records in 28 missing files:  {new_total}')
print(f'Unique NationCodes in 28 files: {len(new_unique_nc)}')
print(f'Overlap with main file: {overlap}')
print(f'Completely new NationCodes: {new_unique_only}')
print(f'---')
print(f'Total records if merged: {existing_total + new_total}')
print(f'Total unique NationCodes: {len(existing_nationcodes | new_unique_nc)}')
