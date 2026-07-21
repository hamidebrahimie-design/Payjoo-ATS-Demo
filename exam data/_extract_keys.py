import openpyxl, os, re

path = r'D:\Payjoo-ATS-Demo\exam data'
missing = ['4102','4177','4179','4180','4181','4182','4183','4184','4185','4186','4187','4188','4189','4190','4191','4192','4193','4194','4195','4196','4197','4198','4199','50013135','50027440','50038035','50047240','50088030']

# Load existing keys from ورود اطلاعات (Key = ExamCode+NationCode)
wb = openpyxl.load_workbook(os.path.join(path, 'ورود اطلاعات.xlsx'), read_only=True, data_only=True)
ws = wb['ثبت نام']
existing_keys = set()
existing_total = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    if any(v is not None for v in row):
        existing_total += 1
        if row[5] is not None:
            existing_keys.add(str(row[5]))
wb.close()

print(f'Existing keys in ورود اطلاعات: {existing_total} records, {len(existing_keys)} unique keys')
print(f'Unique examcodes in اصلی: {len(set())}')
print()

# Extract from each missing file - consider Key as ExamCode(filename)+NationCode(or PersonnelCode)
file_details = []
all_new_keys = set()
grand_total = 0

for code in missing:
    fp = os.path.join(path, f'{code}.xlsx')
    if not os.path.exists(fp):
        fp = os.path.join(path, f'{code}.XLSX')
    if not os.path.exists(fp):
        file_details.append((code, 0, 0, 0, 0))
        continue
    
    wb2 = openpyxl.load_workbook(fp, read_only=True, data_only=True)
    ws2 = wb2['ثبت نام']
    cnt = 0
    keys_this_file = set()
    # determine columns: read header
    headers = {}
    for c in range(1, ws2.max_column+1):
        v = ws2.cell(row=1, column=c).value
        if v:
            headers[c] = str(v).strip()

    # Find which column is the unique identifier (شماره پرسنلي or کد ملی)
    # Check: first column is usually شماره پرسنلي
    id_col = 1  # default: column A = شماره پرسنلي
    
    for row in ws2.iter_rows(min_row=2, values_only=True):
        if any(v is not None for v in row):
            cnt += 1
            person_id = str(row[0]) if row[0] is not None else ''
            key = code + person_id
            keys_this_file.add(key)
            all_new_keys.add(key)
    wb2.close()
    
    new_keys = keys_this_file - existing_keys
    new_keys_cnt = len(new_keys)
    overlap = len(keys_this_file & existing_keys)
    
    file_details.append((code, cnt, len(keys_this_file), overlap, new_keys_cnt))
    grand_total += cnt

print(f'{"ExamCode":<12} {"Records":<9} {"Unique Keys":<12} {"Overlap":<8} {"New Keys":<9}')
print('-' * 50)
for code, rec, ukeys, ovlp, newk in file_details:
    print(f'{code:<12} {rec:<9} {ukeys:<12} {ovlp:<8} {newk:<9}')

total_new_keys = len(all_new_keys - existing_keys)
total_unique_keys = len(all_new_keys | existing_keys)

print()
print('=== Summary ===')
print(f'Current records in ورود اطلاعات: {existing_total}')
print(f'Records in 28 files:             {grand_total}')
print(f'Total if merged:                 {existing_total + grand_total}')
print(f'Unique keys in 28 files:         {len(all_new_keys)}')
print(f'Overlap with existing:           {len(all_new_keys & existing_keys)}')
print(f'Completely new unique keys:      {total_new_keys}')
print(f'Total unique keys (all):         {total_unique_keys}')
