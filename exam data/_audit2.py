import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import pandas as pd
import numpy as np
from openpyxl import load_workbook

f = r'D:\Payjoo-ATS-Demo\exam data\_main.xlsx'
f4101 = r'D:\Payjoo-ATS-Demo\exam data\4101.xlsx'

# Load all sheets once
print("Loading data...")
df_status = pd.read_excel(f, sheet_name='جدول وضعیت')
df_reg = pd.read_excel(f, sheet_name='ثبت نام')
df_kot = pd.read_excel(f, sheet_name='کتبی')
df_mah = pd.read_excel(f, sheet_name='مهارتی')
df_int = pd.read_excel(f, sheet_name='مصاحبه')
df_ac = pd.read_excel(f, sheet_name='کانون')
print("Done loading.\n")

print('='*70)
print('1. مغایرت جدول وضعیت vs داده واقعی')
print('='*70)

status_map = {}
for _, r in df_status.iterrows():
    code = r['کد']
    if pd.isna(code): continue
    code = int(code)
    status_map[code] = r

reg_counts = df_reg['ExamCode'].value_counts().to_dict()
kot_counts = df_kot['ExamCode'].value_counts().to_dict()
mah_counts = df_mah['ExamCode'].value_counts().to_dict()
int_counts = df_int['ExamCode'].value_counts().to_dict()
ac_counts = df_ac['ExamCode'].value_counts().to_dict()

mismatch_count = 0
for code in sorted(status_map.keys()):
    r = status_map[code]
    
    def get_val(c):
        v = r.get(c, np.nan)
        v_str = str(v).strip()
        if v_str == '*' or v_str == 'nan' or v_str == '':
            return None
        try:
            return int(float(v_str))
        except:
            return None

    s_reg = get_val('تعداد ثبت\u200cنام')
    s_elg = get_val('تعداد واجد شرایط')
    s_written_present = get_val('تعداد حاضرین آزمون کتبی')
    s_skill_out = get_val('تعداد نفرات خروجی کتبی')
    s_interview = get_val('تعداد حاضرین مصاحبه')
    s_kanun = get_val('تعداد معرفی به کانون')

    a_reg = reg_counts.get(code, 0)
    a_kot = kot_counts.get(code, 0)
    a_mah = mah_counts.get(code, 0)
    a_int = int_counts.get(code, 0)
    a_ac = ac_counts.get(code, 0)

    issues = []
    if s_reg is not None and a_reg != s_reg:
        issues.append(f'ثبت\u200cنام: جدول={s_reg} واقعی={a_reg}')
    if s_written_present is not None and a_kot != s_written_present:
        issues.append(f'حاضرین کتبی: جدول={s_written_present} واقعی={a_kot}')

    if issues:
        mismatch_count += 1
        nm = r.get('عنوان پست', '')
        print(f'  کد {code} ({str(nm)[:30]}):')
        for iss in issues:
            print(f'    -> {iss}')

print(f'\nمجموع کدهای دارای مغایرت: {mismatch_count}')
print()

print('='*70)
print('2. نمرات منفی یا غیرعادی در کتبی')
print('='*70)
neg_scores = df_kot[df_kot['ScoreW'] < 0]
print(f'  نمرات منفی: {len(neg_scores)}')
if len(neg_scores) > 0:
    print(neg_scores[['ExamCode','NationCode','ScoreW']].to_string())
    
zero_score = df_kot[df_kot['ScoreW'] == 0]
print(f'  نمرات صفر در کتبی: {len(zero_score)}')

print()

print('='*70)
print('3. رکوردهای تکراری کد ملی در یک مرحله')
print('='*70)
for name, df in [('کتبی', df_kot), ('مهارتی', df_mah), ('مصاحبه', df_int), ('کانون', df_ac)]:
    dups = df[df.duplicated(subset=['ExamCode','NationCode'], keep=False)]
    if len(dups) > 0:
        print(f'  {name}: {len(dups)} رکورد تکراری')
        print(f'     کدهای آزمون: {sorted(dups["ExamCode"].unique())}')
    else:
        print(f'  {name}: بدون تکرار')

print()

print('='*70)
print('4. افراد در مرحله پایین‌تر بدون مرحله بالاتر')
print('='*70)
all_codes = set(df_mah['NationCode'].dropna().astype('Int64').astype(str))
kot_ok_codes = set(df_kot[df_kot['Result1']=='مجاز']['NationCode'].dropna().astype('Int64').astype(str))
mah_no_kot = all_codes - kot_ok_codes
print(f'  مهارتی بدون مجاز در کتبی: {len(mah_no_kot)} نفر')

# Interview but no skill
int_codes = set(df_int['NationCode'].dropna().astype('Int64').astype(str))
mah_codes = set(df_mah['NationCode'].dropna().astype('Int64').astype(str))
int_no_mah = int_codes - mah_codes
print(f'  مصاحبه بدون مهارتی: {len(int_no_mah)} نفر')

# AC but no skill
ac_codes = set(df_ac['NationCode'].dropna().astype('Int64').astype(str))
ac_no_mah = ac_codes - mah_codes
print(f'  کانون بدون مهارتی: {len(ac_no_mah)} نفر')

print()

print('='*70)
print('5. شیت قبولی نهایی - آیا رکوردی دارد؟')
print('='*70)
wb = load_workbook(f, read_only=True, data_only=True)
ws = wb['قبولی نهایی']
non_empty = 0
for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
    if row[0] is not None:
        non_empty += 1
    if i > 50000:
        break
wb.close()
print(f'  ردیف‌های غیرخالی در 50000 ردیف اول: {non_empty}')
print(f'  (شیت 1,048,575 ردیف دارد، اکثراً خالی)')

print()

print('='*70)
print('6. مغایرت در 4101.xlsx - ثبت نام')
print('='*70)
df_r_4101 = pd.read_excel(f4101, sheet_name='ثبت نام')
print(f'  4101.xlsx ثبت نام: {len(df_r_4101)} نفر')
print(f'  فایل اصلی: {reg_counts.get(4101, 0)} نفر')
personnel_dups = df_r_4101[df_r_4101.duplicated(subset=['شماره پرسنلي'], keep=False)]
print(f'  شماره پرسنلی تکراری: {len(personnel_dups)}')

print()

print('='*70)
print('7. بررسی Result: افراد نامجاز در مراحل بالاتر')
print('='*70)
# در 4101.xlsx: نتایج کتبی و تحلیل مهارتی
df_written_4101 = pd.read_excel(f4101, sheet_name='نتایج کتبی')
# First row seems to be formula/header row with -1 values
real_rows = df_written_4101[df_written_4101['nacode'] > 0]
print(f'  نتایج کتبی (واقعی): {len(real_rows)} نفر')

# Check if the scores match between main and 4101
main_scores = df_kot[df_kot['ExamCode']==4101][['NationCode','ScoreW']].copy()
main_scores['NationCode'] = main_scores['NationCode'].astype('Int64').astype(str)
main_scores = main_scores.set_index('NationCode')

# 4101 analysis has total written score
df_analysis = pd.read_excel(f4101, sheet_name='تحلیل مهارتی')
analysis_scores = df_analysis[['کد ملی', 'کل کتبی']].copy()
analysis_scores['کد ملی'] = analysis_scores['کد ملی'].dropna().astype('Int64').astype(str)
analysis_scores = analysis_scores.set_index('کد ملی')

# Compare sample
print('  مقایسه نمرات کتبی (نمونه ۵ نفر):')
common = main_scores.index.intersection(analysis_scores.index)
for nc in common[:5]:
    m = main_scores.loc[nc, 'ScoreW']
    a = analysis_scores.loc[nc, 'کل کتبی']
    diff = abs(m - a)
    print(f'    کدملی {nc}: اصلی={m:.2f} / تحلیل={a:.2f} (اختلاف={diff:.2f})')

print()
print('تحلیل کامل شد.')
